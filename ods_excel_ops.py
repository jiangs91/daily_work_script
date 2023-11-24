#代码功能：元数据明细预处理，读取一个有几十万行数据的Excel文件，提取D列包含技术属性的数据

from openpyxl import load_workbook

file_name = 'D:\Code\excel_ops\ods_meta.xlsx'
wb_from = load_workbook(file_name)

ws = wb_from.active
ws1 = wb_from.create_sheet("技术属性")

c = ws['D1']
#print(c.value)

# function:将空格分隔的含有:的字符串转换为字典对象
# input: 技术属性 表名:ods_dd_user_score_advice_full 存储格式:ORC 最近同步时间:2023-11-20 16:23:34 数据库:dw_ods 存储位置:hdfs://nameservice1/inceptor1/user/hive/warehouse/dw_ods.db/hive/ods_dd_user_score_advice_full 存储大小:19.10MB DDL最后变更时间:2023-11-18 03:12:55 表创建 时间:2023-09-06 16:10:42 通用业务属性 负责人:admin@dtstack.com 表中文名:滴滴积分同步申请 个性业务属性
# output: dict['表名'] = 'ods_dd_user_score_advice_full'
def str_to_dict(str):
    sep_list = str.split(' ')
    res = {}
    for item in sep_list:
        if ':' in item:
            item_list = item.split(':')
            res[item_list[0]] = item_list[1]
    return res

#sheet初始化
ws1.cell(row=1, column=1, value='表名')
ws1.cell(row=1, column=2, value='表中文名')
ws1.cell(row=1, column=3, value='存储格式')
ws1.cell(row=1, column=4, value='数据库')
ws1.cell(row=1, column=5, value='存储大小')
ws1.cell(row=1, column=6, value='表行数')

colD = ws['D']
count = 2

for cell in colD:
    
    if cell.value is not None and cell.value[:4] == '技术属性':
        print(cell.value)
        dict_item = str_to_dict(cell.value)
        ws1.cell(row=count, column=1, value = dict_item.get('表名', 'None'))
        ws1.cell(row=count, column=2, value = dict_item.get('表中文名', 'None'))
        ws1.cell(row=count, column=3, value = dict_item.get('存储格式', 'None'))
        ws1.cell(row=count, column=4, value = dict_item.get('数据库', 'None'))
        ws1.cell(row=count, column=5, value = dict_item.get('存储大小', 'None'))
        ws1.cell(row=count, column=6, value = dict_item.get('表行数', 'None'))
        count = count + 1
    
    if cell.value == 'EndOfAll':
        break

wb_from.save(file_name)

